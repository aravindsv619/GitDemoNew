from TestData.excelDemo import Dict


class HomePageData:
    test_HomePage_data = [{"firstname": "Aravind", "email": "aravindsv619@gmail.com", "gender": "Male"},
                          {"firstname": "Mukund", "email": "mukundsv333@gmail.com", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):

        import openpyxl
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\aravi\\Desktop\\Selenium Python Automation\\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
