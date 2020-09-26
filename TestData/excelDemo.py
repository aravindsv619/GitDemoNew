import openpyxl

book = openpyxl.load_workbook("C:\\Users\\aravi\\Desktop\\Selenium Python Automation\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
